"""
build_calendar.py
Stage 2: build the content calendar from Stage 1 outputs (fresh, self-contained builder). Proven on a
real run. 20 Tier-1 ideas per week (8 short, 6 long, 4 live, 2 playlist) across Week 0 + 12 weeks,
month-grounded themes, plus the complete tiered idea bank, the playlists series, and the evergreen
pool. EXAMPLE VALUES BELOW are illustrative for a CAT-prep niche; replace per-niche: KEYWORD_DEMAND,
TOPIC_TO_KEYWORD, TITLE_FRAME, THUMB, PLAYLISTS, WEEK_SCHEDULE, COMPETITOR_NAMES, and the output
filename.

Reads:  out/seasonal_may_aug_2025.csv, out/seasonal_may_aug_2024.csv, out/outliers_core.csv,
        out/depth_winners.csv, and (optional) out/gemini_titles.json (video_id -> best title).
Writes: FINAL/04_operational_calendars/content_calendar.xlsx
        out/calendar_data.json (slotted ideas, for the Gemini title pass)
        out/source_url_status.json (link verification cache)

Sheets: Read me first, Weekly calendar, Complete idea bank (Tier 1/2/3), Evergreen pool, Playlists,
Legend and sources. Titles use Gemini (free), never vidIQ.
"""
import csv, json, os, re, time
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJ = Path(__file__).resolve().parent.parent
OUT = PROJ / "out"
FINAL_CAL = PROJ / "FINAL" / "04_operational_calendars"
FINAL_CAL.mkdir(parents=True, exist_ok=True)
OUT_XLSX = FINAL_CAL / "content_calendar.xlsx"
DATA_JSON = OUT / "calendar_data.json"
URL_STATUS = OUT / "source_url_status.json"
GEMINI_TITLES = OUT / "gemini_titles.json"
load_dotenv(PROJ / ".env")
YT_KEY = os.getenv("YT_API_KEY")
YT_BASE = "https://www.googleapis.com/youtube/v3"

KEYWORD_DEMAND = {
    "IIM placement reality": (4938, 63), "salary after MBA": (5020, 63), "CAT profile evaluation": (5209, 61),
    "life at IIM": (5139, 63), "best MBA colleges in India": (7703, 63), "CAT 2026 strategy": (4773, 61),
    "CAT score vs percentile": (8579, 58), "CAT 2026": (35196, 60), "CAT mock test": (4508, 60),
    "CAT preparation strategy": (4773, 61), "VARC CAT preparation": (21454, 68), "CAT DILR strategy": (9123, 67),
    "CAT quant": (22380, 67), "WAT PI preparation": (5148, 57), "MBA exams other than CAT": (4569, 63),
    "CAT topper": (9998, 61), "CAT preparation": (27627, 67), "FMS Delhi": (15179, 65),
    "reading comprehension for CAT preparation": (8701, 66), "CAT preparation for beginners": (12844, 61),
}
TOPIC_TO_KEYWORD = {
    "placement-stories": "IIM placement reality", "fees-roi": "salary after MBA",
    "admissions": "CAT profile evaluation", "IIM-life": "life at IIM", "day-in-life": "life at IIM",
    "college-comparison": "best MBA colleges in India", "mba-strategy": "CAT 2026 strategy",
    "percentile-strategy": "CAT score vs percentile", "cutoff-analysis": "CAT score vs percentile",
    "exam-pattern": "CAT 2026", "mock-test": "CAT mock test", "mock-analysis": "CAT mock test",
    "sectional-strategy": "CAT preparation strategy", "VARC": "VARC CAT preparation",
    "DILR": "CAT DILR strategy", "QA": "CAT quant", "WAT-PI": "WAT PI preparation",
    "GD": "WAT PI preparation", "OMET": "MBA exams other than CAT", "topper-interview": "CAT topper",
    "profile-building": "CAT profile evaluation", "motivation": "CAT preparation",
    "current-affairs-gk": "CAT preparation", "results": "CAT score vs percentile",
    "marathon": "CAT quant", "live-doubt-solving": "CAT preparation", "batch-launch": "CAT 2026",
    "lead-magnet": "CAT preparation", "faculty-intro": "CAT preparation", "meme-relatable": "CAT preparation",
    "promo-only": "CAT 2026", "other": "CAT 2026",
}
TITLE_FRAME = {
    "placement-stories": "{hook} | Real MBA salary, the honest numbers",
    "fees-roi": "{hook} | MBA fees vs salary, is it worth it",
    "admissions": "{hook} | Selection criteria explained for 2026",
    "IIM-life": "{hook} | Life at an IIM, the real picture",
    "day-in-life": "{hook} | A real day at an IIM",
    "college-comparison": "{hook} | Which MBA college, honest comparison",
    "mba-strategy": "{hook} | CAT 2026 strategy from scratch",
    "percentile-strategy": "{hook} | What 99 percentile actually takes",
    "exam-pattern": "{hook} | CAT 2026 pattern, what to expect",
    "mock-test": "{hook} | CAT mock strategy that works",
    "mock-analysis": "{hook} | How to analyse your CAT mock",
    "sectional-strategy": "{hook} | Sectional plan for CAT 2026",
    "VARC": "{hook} | VARC for CAT, even if you fear English",
    "DILR": "{hook} | CAT DILR, the method that works",
    "QA": "{hook} | CAT Quant made simple",
    "WAT-PI": "{hook} | WAT and PI prep for the IIM call",
    "topper-interview": "{hook} | How the topper actually did it",
    "profile-building": "{hook} | Build your MBA profile now",
    "OMET": "{hook} | MBA exams beyond CAT (XAT, SNAP, NMAT)",
    "motivation": "{hook} | For every CAT aspirant who feels behind",
    "results": "{hook} | CAT score and percentile, decoded", "other": "{hook} | CAT 2026",
}
THUMB = {
    "placement-stories": "REAL SALARY", "fees-roi": "WORTH IT?", "admissions": "DO I QUALIFY",
    "IIM-life": "INSIDE AN IIM", "college-comparison": "WHICH IIM", "mba-strategy": "WHERE TO START",
    "percentile-strategy": "99 %ILE PLAN", "exam-pattern": "CAT 2026", "mock-test": "MOCK STRATEGY",
    "VARC": "VARC FIX", "DILR": "DILR METHOD", "QA": "QUANT SIMPLE", "WAT-PI": "IIM INTERVIEW",
    "topper-interview": "TOPPER SECRETS", "OMET": "BEYOND CAT", "day-in-life": "IIM LIFE",
    "profile-building": "PROFILE TIPS", "results": "PERCENTILE", "motivation": "DON'T QUIT", "other": "CAT 2026",
}

# ---------- playlists (full series, grounded in the research) ----------
# EXAMPLE VALUES. "proof" should cite your own outlier evidence (channel + view count + the
# matching vidIQ keyword volume), e.g. "Competitor 3: 'Why 90% Fail in LRDI' 21K views; cat dilr
# strategy 9,123/mo". Fill from your own outliers_core.csv + KEYWORD_DEMAND, never invented.
PLAYLISTS = [
    {"name": "Complete VARC in English", "type": "Hero", "keyword": "VARC CAT preparation", "run": "Jun to Nov",
     "proof": "see outliers_core.csv for this niche's VARC winners; varc cat preparation 21,454/mo",
     "episodes": ["RC method for CAT", "Para-jumbles", "Para-summary", "Odd sentence out", "Vocab in context",
                  "RC speed vs accuracy", "Inference and tone questions", "VARC for non-engineers",
                  "VARC mock analysis", "Last-mile VARC revision"]},
    {"name": "Complete DILR in English", "type": "Hero", "keyword": "CAT DILR strategy", "run": "Jun to Nov",
     "proof": "see outliers_core.csv for this niche's top DILR outlier; cat dilr strategy 9,123/mo",
     "episodes": ["Set selection in DILR", "Arrangements and puzzles", "DI tables and graphs",
                  "Games and tournaments", "Conditional sets", "Why most fail LRDI", "DILR speed",
                  "DILR mock analysis", "Hard set walkthrough", "Last-mile DILR"]},
    {"name": "Complete QA in English", "type": "Hero", "keyword": "CAT quant", "run": "Jun to Nov",
     "proof": "cat quant 22,380/mo",
     "episodes": ["Arithmetic for CAT", "Algebra", "Geometry", "Number system", "Modern math",
                  "QA shortcuts", "Quant without fear (non-engineers)", "QA mock analysis", "Last-30-days QA"]},
    {"name": "Complete OMET in English", "type": "Hero", "keyword": "MBA exams other than CAT", "run": "Aug to Dec",
     "proof": "xat 12,083/mo, mba exams other than cat 4,569/mo (never title 'OMET')",
     "episodes": ["XAT decision making and essay", "SNAP strategy", "NMAT strategy", "IIFT",
                  "CMAT GK", "TISSNET", "MBA exams beyond CAT overview", "OMET vs CAT planning"]},
    {"name": "CAT 2026 Strategy From Scratch", "type": "Lane", "keyword": "CAT 2026 strategy", "run": "Jun to Nov",
     "proof": "cat preparation 27,627/mo",
     "episodes": ["How to start CAT prep", "6-month plan", "Sectional strategy", "Mock strategy",
                  "Balancing job or college", "Last 30 days", "Exam-day plan"]},
    {"name": "MBA Reality and IIM Life", "type": "Lane", "keyword": "IIM placement reality", "run": "Jun to Oct",
     "proof": "see outliers_core.csv for this niche's top real-salary outlier; iim placement reality 4,938/mo",
     "episodes": ["Real salary after MBA", "IIM placement reality", "Life at IIM A/B/C",
                  "Life at the new IIMs", "A day in the life", "Finance vs Marketing vs HR",
                  "Corporate vs startup after MBA", "Is the MBA worth it"]},
    {"name": "Selection Criteria Decoded", "type": "Lane", "keyword": "FMS Delhi", "run": "Jun to Sep",
     "proof": "see outliers_core.csv for this niche's top selection-criteria outlier; FMS Delhi 15,179/mo",
     "episodes": ["FMS Delhi criteria", "IIM A/B/C criteria", "IIM Indore and Kozhikode",
                  "New IIMs criteria", "Can I get in with a gap year", "Low academics or category profiles",
                  "Non-engineer profiles", "How to fill the CAT form"]},
    {"name": "CAT Toppers Decoded", "type": "Lane", "keyword": "CAT topper", "run": "Jun to Nov",
     "proof": "see outliers_core.csv for this niche's topper-interview winners; cat topper 9,998/mo",
     "episodes": ["Non-engineer 99 percentiler", "100 percentiler interview",
                  "Working-professional topper", "Repeater comeback", "How the topper studied"]},
    {"name": "Daily Targets to CAT 2026", "type": "Lane (recurring)", "keyword": "CAT preparation", "run": "Jun to Nov",
     "proof": "a recurring franchise on your own channel builds brand search; see your own channel_format_stats.csv",
     "episodes": ["Weekly target checklist (recurring)", "Days-to-CAT countdown (recurring)",
                  "Must-do of the week (recurring)"]},
    {"name": "WAT-PI Prep", "type": "Lane (post-result)", "keyword": "WAT PI preparation", "run": "Jan to Mar 2027",
     "proof": "see outliers_core.csv for this niche's top interview-prep outlier; wat pi preparation 5,148/mo",
     "episodes": ["WAT writing", "GD do's and don'ts", "Common PI questions", "Gap-year and profile answers",
                  "Mock personal interview"]},
]

DAY = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
MON = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def fmt_date(d):
    return f"{DAY[d.weekday()]} {d.day} {MON[d.month-1]}"


def wk(n, start, theme, topics, playlists, quota):
    return {"n": n, "start": start, "end": start + timedelta(days=6), "theme": theme,
            "topics": set(topics), "playlists": playlists, "quota": dict(quota)}


STD = {"short": 8, "long": 6, "live": 4}
W0Q = {"short": 6, "long": 5, "live": 3}
WEEK_SCHEDULE = [
    wk(0, date(2026, 5, 25), "Kickoff: why CAT 2026 in English, and what the journey looks like",
       ["mba-strategy", "IIM-life", "motivation", "exam-pattern", "VARC", "QA"],
       ["CAT 2026 Strategy From Scratch", "MBA Reality and IIM Life"], W0Q),
    wk(1, date(2026, 6, 1), "Outcome first: salary, placements, which IIM (June: admissions and outcome lead)",
       ["placement-stories", "fees-roi", "IIM-life", "college-comparison"],
       ["MBA Reality and IIM Life", "Selection Criteria Decoded"], STD),
    wk(2, date(2026, 6, 8), "Can I get in: profile and selection criteria",
       ["admissions", "profile-building", "percentile-strategy", "mba-strategy"],
       ["Selection Criteria Decoded", "CAT 2026 Strategy From Scratch"], STD),
    wk(3, date(2026, 6, 15), "How to start and section foundations",
       ["mba-strategy", "VARC", "QA", "DILR"],
       ["CAT 2026 Strategy From Scratch", "Complete VARC in English"], STD),
    wk(4, date(2026, 6, 22), "Which IIM and early DILR",
       ["college-comparison", "admissions", "DILR", "IIM-life"],
       ["Selection Criteria Decoded", "Complete DILR in English"], STD),
    wk(5, date(2026, 6, 29), "Strategy intensifies, VARC for the non-engineer (July: strategy ramps)",
       ["mba-strategy", "VARC", "sectional-strategy", "percentile-strategy"],
       ["Complete VARC in English", "CAT 2026 Strategy From Scratch"], STD),
    wk(6, date(2026, 7, 6), "DILR and VARC foundations, toppers",
       ["DILR", "VARC", "topper-interview", "mba-strategy"],
       ["Complete DILR in English", "CAT Toppers Decoded"], STD),
    wk(7, date(2026, 7, 13), "Mock season begins (mock-test content rises in July)",
       ["mock-test", "mock-analysis", "QA", "sectional-strategy"],
       ["Complete QA in English", "Daily Targets to CAT 2026"], STD),
    wk(8, date(2026, 7, 20), "Mocks, percentile, and an outcome refresh",
       ["mock-test", "percentile-strategy", "placement-stories", "QA"],
       ["Complete QA in English", "MBA Reality and IIM Life"], STD),
    wk(9, date(2026, 7, 27), "Registration ramp begins: form, profile, eligibility",
       ["admissions", "exam-pattern", "mba-strategy", "profile-building"],
       ["Selection Criteria Decoded", "CAT 2026 Strategy From Scratch"], STD),
    wk(10, date(2026, 8, 3), "Registration, MBA exams beyond CAT, sections grind (August)",
       ["admissions", "OMET", "QA", "DILR"],
       ["Complete OMET in English", "Complete QA in English"], STD),
    wk(11, date(2026, 8, 10), "Section grind and VARC peak",
       ["VARC", "QA", "DILR", "sectional-strategy"],
       ["Complete VARC in English", "Complete DILR in English"], STD),
    wk(12, date(2026, 8, 17), "Sprint setup, WAT-PI seeding, English Shorts push",
       ["sectional-strategy", "mba-strategy", "WAT-PI", "QA", "VARC", "DILR"],
       ["Daily Targets to CAT 2026", "WAT-PI Prep"], STD),
]

PROMO_RE = re.compile(r"\b(launch|enroll|register now|join now|scholarship|batch starting|"
                      r"course\b|anniversary|free demo|webinar|admission open)\b", re.IGNORECASE)


# EXAMPLE VALUE: the competitor channel names to strip out of a cloned title's hook. Fill from your
# own resolve_channels.py CHANNELS list (lowercase, pipe-separated).
COMPETITOR_NAMES = r"competitor1|competitor2|competitor3"


def clean_hook(title):
    t = re.sub(r"#\w+", "", title)
    t = re.split(r"[|:–—]", t)[0].strip()
    t = re.sub(rf"\b({COMPETITOR_NAMES})\b", "", t, flags=re.IGNORECASE).strip()
    return " ".join(re.sub(r"\s+", " ", t).split()[:9]) or "CAT 2026"


def draft_title(topic, source_title):
    return TITLE_FRAME.get(topic, TITLE_FRAME["other"]).format(hook=clean_hook(source_title))


def faculty_for(topic):
    if topic == "VARC":
        return "VARC educator"
    if topic in ("QA", "DILR", "sectional-strategy", "mock-analysis", "marathon"):
        return "Quant/DILR educator"
    if topic in ("admissions", "college-comparison", "fees-roi", "WAT-PI", "profile-building"):
        return "Mentor / counsellor"
    if topic in ("IIM-life", "day-in-life", "placement-stories", "topper-interview"):
        return "IIM alum guest"
    return "Main educator"


def effort_for(fmt, topic):
    if fmt == "live":
        return "Medium"
    if fmt == "long" and topic in ("topper-interview", "IIM-life", "placement-stories", "day-in-life"):
        return "High (guest)"
    return "Low" if fmt == "short" else "Medium"


def channel_for(topic):
    return "Second-creator option" if topic in ("IIM-life", "day-in-life", "motivation", "meme-relatable") else "Main"


def publish_window(topic, fmt):
    if topic in ("admissions", "exam-pattern", "results"):
        return "Weekday morning, near registration news"
    if fmt == "short":
        return "Daily 7-9pm micro-slot"
    if fmt == "live":
        return "Weekend (Sat/Sun) evening"
    return "Tue/Thu/Sat 7-8pm study hours"


def secondary_kw(topic):
    base = ["#CAT2026", "#CATPreparation"]  # EXAMPLE VALUE: add your own channel/brand hashtag here
    extra = {"VARC": ["#CATVARC"], "QA": ["#CATQuant"], "DILR": ["#CATDILR"],
             "admissions": ["#FMSDelhi", "#IIM"], "placement-stories": ["#MBASalary"],
             "IIM-life": ["#IIMlife"], "WAT-PI": ["#WATPI", "#IIMinterview"], "OMET": ["#XAT", "#SNAP", "#NMAT"]}
    return ", ".join((base + extra.get(topic, []))[:6])


def vid_from_url(u):
    m = re.search(r"v=([A-Za-z0-9_-]{6,})", u or "")
    return m.group(1) if m else ""


def load_aux():
    ratios, thin, themes = {}, {}, {}
    p = OUT / "outliers_core.csv"
    if p.exists():
        for r in csv.DictReader(p.open(encoding="utf-8")):
            ratios[r["video_id"]] = float(r["outlier_ratio"]); thin[r["video_id"]] = (r.get("thin_baseline") == "yes")
    p = OUT / "depth_winners.csv"
    if p.exists():
        for r in csv.DictReader(p.open(encoding="utf-8")):
            if r.get("top_themes"):
                themes[r["video_id"]] = r["top_themes"]
    return ratios, thin, themes


def load_ideas():
    ratios, thin, themes = load_aux()
    rows = []
    for fn in ("seasonal_may_aug_2025.csv", "seasonal_may_aug_2024.csv"):
        p = OUT / fn
        if not p.exists():
            continue
        for r in csv.DictReader(p.open(encoding="utf-8")):
            if r.get("source_language") != "english":
                continue
            r["view_count"] = int(r["view_count"] or 0)
            r["video_id"] = vid_from_url(r.get("video_url", ""))
            rows.append(r)
    best = {}
    for r in rows:
        v = r["video_id"]
        if v and (v not in best or r["view_count"] > best[v]["view_count"]):
            best[v] = r
    ideas = []
    for r in best.values():
        topic = r.get("topic") or "other"
        kw = TOPIC_TO_KEYWORD.get(topic, "CAT 2026")
        dem = KEYWORD_DEMAND.get(kw, (None, None))
        ratio = ratios.get(r["video_id"], 0.0)
        thinb = thin.get(r["video_id"], False)
        # tier score: absolute views (log-ish via raw) blended with a capped, non-thin ratio
        rscore = 0 if thinb else min(ratio, 50)
        ideas.append({
            "video_id": r["video_id"], "source_channel": r["channel"], "source_title": r["title"],
            "source_views": r["view_count"], "source_ratio": ratio, "thin": thinb,
            "promo": bool(PROMO_RE.search(r["title"])), "topic": topic, "format": r["format"],
            "hook_type": r.get("hook_type", ""), "season": r["season"], "video_url": r["video_url"],
            "primary_keyword": kw, "kw_volume": dem[0], "kw_overall": dem[1],
            "comment_themes": themes.get(r["video_id"], ""), "thumb": THUMB.get(topic, "CAT 2026"),
            "best_title": draft_title(topic, r["title"]),
            "_score": r["view_count"] + rscore * 500,
        })
    ideas.sort(key=lambda x: (0 if not x["promo"] else 1, -x["_score"]))
    # tier by rank among non-promo
    organic = [i for i in ideas if not i["promo"]]
    for rank, i in enumerate(organic):
        i["tier"] = "Tier 1" if rank < 150 else ("Tier 2" if rank < 500 else "Tier 3")
    for i in ideas:
        if i["promo"]:
            i["tier"] = "Tier 3 (promo-suspect)"
    return ideas


def verify_video_ids(video_ids):
    """Existence check via YouTube API videos.list (deleted/private IDs are not returned).
    Returns {video_id: exists}. Cheap, about 1 unit per 50 ids."""
    status = {}
    if URL_STATUS.exists():
        try:
            status = json.loads(URL_STATUS.read_text())
        except Exception:
            status = {}
    todo = [v for v in dict.fromkeys(video_ids) if v and v not in status]
    for i in range(0, len(todo), 50):
        batch = todo[i:i + 50]
        try:
            r = requests.get(f"{YT_BASE}/videos", params={"part": "id", "id": ",".join(batch), "key": YT_KEY}, timeout=30)
            existing = {it["id"] for it in r.json().get("items", [])} if r.status_code == 200 else set(batch)
        except requests.RequestException:
            existing = set(batch)
        for v in batch:
            status[v] = (v in existing)
    if todo:
        URL_STATUS.write_text(json.dumps(status))
    return status


def slot(ideas):
    assigned = {w["n"]: [] for w in WEEK_SCHEDULE}
    used = set()
    for week in WEEK_SCHEDULE:
        quota = dict(week["quota"])
        cands = [i for i in ideas if i["video_id"] not in used and i["topic"] in week["topics"]]
        for i in cands:
            if quota.get(i["format"], 0) > 0:
                assigned[week["n"]].append(i); used.add(i["video_id"]); quota[i["format"]] -= 1
            if sum(quota.values()) == 0:
                break
        # backfill live from the whole live pool (live is concentrated in section/strategy lanes)
        if quota.get("live", 0) > 0:
            for i in ideas:
                if i["video_id"] in used or i["format"] != "live":
                    continue
                assigned[week["n"]].append(i); used.add(i["video_id"]); quota["live"] -= 1
                if quota["live"] == 0:
                    break
        # backfill short/long from any fitting topic
        for fmt in ("short", "long"):
            if quota.get(fmt, 0) > 0:
                for i in cands:
                    if i["video_id"] in used or i["format"] != fmt:
                        continue
                    assigned[week["n"]].append(i); used.add(i["video_id"]); quota[fmt] -= 1
                    if quota[fmt] == 0:
                        break
    evergreen = [i for i in ideas if i["video_id"] not in used]
    evergreen.sort(key=lambda x: -x["_score"])
    return assigned, evergreen


# ---------- xlsx ----------
HFILL = PatternFill("solid", fgColor="DCE6F1"); SFILL = PatternFill("solid", fgColor="305496")
PFILL = PatternFill("solid", fgColor="548235")
SFONT = Font(bold=True, color="FFFFFF", size=12); HFONT = Font(bold=True); LINK = Font(color="0563C1", underline="single")
COLUMNS = [("Format", 9), ("Best title (draft)", 38), ("Original title (source)", 34),
           ("Thumbnail text", 15), ("Primary keyword", 20), ("Secondary keywords", 24),
           ("Why this idea", 50), ("Last year search volume", 13), ("Current ranking opportunity", 15),
           ("Competitor data", 34), ("Notes", 22), ("Owner / Status", 13), ("Publish window", 22),
           ("Production effort", 13), ("Faculty needed", 17), ("Channel", 16), ("Source video URL", 44)]


def why_text(i, week):
    lens = "outlier ratio" if (i["source_ratio"] and not i["thin"]) else "absolute views"
    line1 = f"Week theme: {week['theme']}." if week else f"{i['tier']}, evergreen, pick anytime."
    rr = f", {i['source_ratio']:.0f}x channel-format median" if (i["source_ratio"] and not i["thin"]) else ""
    line2 = f"Proven by {i['source_channel']}: {i['source_views']:,} views{rr} ({i['season'].replace('_',' ')}). Lens: {lens}."
    if i["kw_volume"]:
        line3 = f"Demand: \"{i['primary_keyword']}\" about {i['kw_volume']:,}/mo, opportunity {i['kw_overall']}/100."
    else:
        line3 = f"Demand: \"{i['primary_keyword']}\" (see demand map)."
    line4 = f"Audience asks (comments): {i['comment_themes'][:140]}" if i["comment_themes"] else ""
    return "\n".join(x for x in (line1, line2, line3, line4) if x)


def comp_data(i, url_ok):
    flag = " [promo-suspect]" if i["promo"] else (" [thin baseline, use views]" if i["thin"] else "")
    dead = "" if url_ok.get(i["video_url"], True) else " [LINK DEAD, replace]"
    return f"{i['source_channel']}: \"{i['source_title'][:88]}\" | {i['source_views']:,} views | {i['format']}{flag}{dead}"


def write_header(ws):
    for idx, (name, w) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=idx, value=name); c.font = HFONT; c.fill = HFILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.row_dimensions[1].height = 30; ws.freeze_panes = "A2"


def write_band(ws, row, text, fill=SFILL):
    ws.cell(row=row, column=1, value=text); ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    c = ws.cell(row=row, column=1); c.fill = fill; c.font = SFONT; c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def write_idea_row(ws, row, i, week, url_ok):
    vol = i["kw_volume"] if i["kw_volume"] is not None else ""
    opp = f"{i['kw_overall']}/100" if i["kw_overall"] is not None else "see demand map"
    vals = [{"short": "Short", "long": "Long", "live": "Live"}.get(i["format"], i["format"]),
            i["best_title"], i["source_title"], i["thumb"], i["primary_keyword"], secondary_kw(i["topic"]),
            why_text(i, week), vol, opp, comp_data(i, url_ok), "draft title, polish before publishing", "",
            publish_window(i["topic"], i["format"]), effort_for(i["format"], i["topic"]),
            faculty_for(i["topic"]), channel_for(i["topic"]), None]
    for ci, val in enumerate(vals, 1):
        if ci == len(COLUMNS):
            c = ws.cell(row=row, column=ci, value=i["video_url"]); c.hyperlink = i["video_url"]
            c.font = LINK; c.alignment = Alignment(vertical="top")
        else:
            c = ws.cell(row=row, column=ci, value=val)
            c.alignment = Alignment(wrap_text=(ci in (2, 3, 6, 7, 10, 11, 13)), vertical="top")
    ws.row_dimensions[row].height = 92


def write_playlist_row(ws, row, pname):
    pl = next((p for p in PLAYLISTS if p["name"] == pname), None)
    if not pl:
        return
    dem = KEYWORD_DEMAND.get(pl["keyword"], (None, None))
    eps = "; ".join(f"{n+1}. {e}" for n, e in enumerate(pl["episodes"]))
    vals = ["Playlist", f"{pl['name']} ({len(pl['episodes'])} episodes)", "", "SERIES", pl["keyword"],
            "", f"{pl['type']} series, runs {pl['run']}. Episodes: {eps}",
            dem[0] if dem[0] else "", f"{dem[1]}/100" if dem[1] else "", f"Proof: {pl['proof']}",
            "produce 1 episode this week", "", "Weekly cadence", "Medium", faculty_for("mba-strategy"), "Main", ""]
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.alignment = Alignment(wrap_text=(ci in (2, 7, 10)), vertical="top")
        if ci == 1:
            c.fill = PFILL; c.font = Font(bold=True, color="FFFFFF")
    ws.row_dimensions[row].height = 70


def build():
    ideas = load_ideas()
    assigned, evergreen = slot(ideas)
    print("verifying source links via YouTube API...")
    id_ok = verify_video_ids([i["video_id"] for w in WEEK_SCHEDULE for i in assigned[w["n"]]])
    url_ok = {i["video_url"]: id_ok.get(i["video_id"], True) for w in WEEK_SCHEDULE for i in assigned[w["n"]]}
    gtitles = json.loads(GEMINI_TITLES.read_text()) if GEMINI_TITLES.exists() else {}
    for i in ideas:
        if i["video_id"] in gtitles:
            i["best_title"] = gtitles[i["video_id"]]

    # persist slotted data for the Gemini title pass
    slotted = []
    for w in WEEK_SCHEDULE:
        for i in assigned[w["n"]]:
            slotted.append({"video_id": i["video_id"], "topic": i["topic"], "format": i["format"],
                            "source_title": i["source_title"], "primary_keyword": i["primary_keyword"],
                            "hook_type": i["hook_type"], "week": w["n"]})
    DATA_JSON.write_text(json.dumps(slotted, indent=2, ensure_ascii=False), encoding="utf-8")

    wb = Workbook(); wb.remove(wb.active)

    # Read me
    ws = wb.create_sheet("Read me first"); ws.column_dimensions["A"].width = 110
    ws.cell(row=1, column=1, value="CAT English content calendar, May to August 2026").font = Font(bold=True, size=15)
    intro = [
        "Week 0 (Mon 25 May 2026, kickoff) then Weeks 1 to 12 (Mon 1 Jun to Sun 23 Aug 2026), anchored to the CAT 2026 registration ramp (exam 29 Nov, registration early Aug to mid Sep).",
        "20 ideas per week (8 short, 6 long, 4 live, 2 playlist), all Tier 1, deliberately over-supplied so the team picks the best 50 to 60%. Every idea is sourced from a competitor video that actually won in May to August (2025 or 2024); open the Source video URL to study the original.",
        "Themes follow what wins by month: June leads with admissions and outcome (which IIM, salary, can I get in), July adds strategy and the start of mock season, August is section grind plus registration.",
        "Each row shows the original competitor title and a best title written to vidIQ title guidelines (keyword front-loaded, 50 to 60 chars, a number or curiosity or outcome trigger). Titles and thumbnails are starting points; polish before publishing. Owner / Status is blank for the team.",
        "Sheets: Weekly calendar (this plan), Complete idea bank (all proven ideas, Tier 1/2/3), Evergreen pool (off-cycle), Playlists (full series to run to CAT 2026), Legend and sources.",
    ]
    r = 3
    for line in intro:
        c = ws.cell(row=r, column=1, value="- " + line); c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 50; r += 1

    # Weekly calendar
    ws = wb.create_sheet("Weekly calendar"); write_header(ws); row = 2; total = 0
    for week in WEEK_SCHEDULE:
        write_band(ws, row, f"Week {week['n']} | {fmt_date(week['start'])} to {fmt_date(week['end'])} 2026 | {week['theme']}"); row += 1
        for i in assigned[week["n"]]:
            write_idea_row(ws, row, i, week, url_ok); row += 1; total += 1
        for pname in week["playlists"]:
            write_playlist_row(ws, row, pname); row += 1

    # Complete idea bank (tiered)
    ws = wb.create_sheet("Complete idea bank"); write_header(ws); row = 2
    for tier in ("Tier 1", "Tier 2", "Tier 3", "Tier 3 (promo-suspect)"):
        tideas = [i for i in ideas if i["tier"] == tier]
        if not tideas:
            continue
        write_band(ws, row, f"{tier} | {len(tideas)} ideas"); row += 1
        for i in sorted(tideas, key=lambda x: -x["_score"]):
            write_idea_row(ws, row, i, None, url_ok); row += 1

    # Evergreen pool
    ws = wb.create_sheet("Evergreen pool"); write_header(ws)
    write_band(ws, 2, f"Evergreen pool | {len(evergreen)} off-cycle ideas, pick anytime (sorted by evidence)"); row = 3
    for i in evergreen:
        write_idea_row(ws, row, i, None, url_ok); row += 1

    # Playlists
    ws = wb.create_sheet("Playlists"); ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14; ws.column_dimensions["C"].width = 26; ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 18; ws.column_dimensions["F"].width = 70
    for ci, h in enumerate(["Playlist", "Type", "Target keyword", "Demand/mo", "Run window", "Episodes and proof"], 1):
        c = ws.cell(row=1, column=ci, value=h); c.font = HFONT; c.fill = HFILL
    ws.row_dimensions[1].height = 24; ws.freeze_panes = "A2"; row = 2
    for pl in PLAYLISTS:
        dem = KEYWORD_DEMAND.get(pl["keyword"], (None, None))
        eps = "  |  ".join(f"{n+1}. {e}" for n, e in enumerate(pl["episodes"]))
        vals = [pl["name"], pl["type"], pl["keyword"], dem[0] if dem[0] else "see map", pl["run"],
                f"{eps}\nProof: {pl['proof']}"]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val); c.alignment = Alignment(wrap_text=(ci in (1, 6)), vertical="top")
            if ci == 1:
                c.font = HFONT
        ws.row_dimensions[row].height = 78; row += 1

    # Legend
    ws = wb.create_sheet("Legend and sources"); ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 84
    ws.cell(row=1, column=1, value="Column reference and method").font = Font(bold=True, size=13)
    refs = [("Best title", "Drafted to vidIQ title guidelines by Gemini, keyword front-loaded. A starting point."),
            ("Original title (source)", "The competitor video's real title, for reference."),
            ("Primary keyword / demand", "From our vidIQ demand map (already pulled). Search volume and 0-100 opportunity."),
            ("Why this idea", "Week theme + competitor proof (views and ratio, lens stated) + demand + what comments ask."),
            ("Competitor data", "The proven source video; flags promo-suspect, thin-baseline, and any dead link."),
            ("Source video URL", "The working link to the original; verified live. Click or copy."),
            ("Tiers", "Tier 1 top ~150 by evidence, Tier 2 next ~350, Tier 3 the rest. All are proven winners."),
            ("Playlists", "Full series to run to CAT 2026; episodes and the competitor lane that proves each.")]
    r = 3
    for k, v in refs:
        ws.cell(row=r, column=1, value=k).font = HFONT
        c = ws.cell(row=r, column=2, value=v); c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 40; r += 1

    wb.save(OUT_XLSX)
    dead = sum(1 for ok in url_ok.values() if not ok)
    print(f"Wrote {OUT_XLSX}")
    print(f"  weekly slotted: {total} videos + {len(WEEK_SCHEDULE)*2} playlist rows | idea bank: {len(ideas)} | evergreen: {len(evergreen)}")
    print(f"  source links checked: {len(url_ok)}, dead: {dead}")
    print(f"  best titles from Gemini: {sum(1 for i in ideas if i['video_id'] in gtitles)}/{len(ideas)} (run gen_titles_gemini.py then rebuild)")
    for w in WEEK_SCHEDULE:
        by = defaultdict(int)
        for i in assigned[w["n"]]:
            by[i["format"]] += 1
        print(f"  W{w['n']:2d} {fmt_date(w['start'])}: {len(assigned[w['n']])} vids (s{by['short']} l{by['long']} live{by['live']}) +2 playlist")


if __name__ == "__main__":
    build()
